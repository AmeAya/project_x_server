import openpyxl
from rest_framework.views import APIView


def download_exe(query):
    wb = openpyxl.Workbook()
    wb.create_sheet(title='Sheet1', index=0)
    sheet = wb['Sheet1']

    j = 1
    cell = sheet.cell(row=j, column=1)
    cell.value = 'Дата'
    cell = sheet.cell(row=j, column=2)
    cell.value = 'Материал'
    cell = sheet.cell(row=j, column=3)
    cell.value = 'Марка'
    cell = sheet.cell(row=j, column=4)
    cell.value = 'Объект'
    cell = sheet.cell(row=j, column=5)
    cell.value = 'Блок'
    cell = sheet.cell(row=j, column=6)
    cell.value = 'Этаж'
    cell = sheet.cell(row=j, column=7)
    cell.value = 'Конструктив'
    cell = sheet.cell(row=j, column=8)
    cell.value = 'Единицы измерения'
    cell = sheet.cell(row=j, column=9)
    cell.value = 'Колличество'
    cell = sheet.cell(row=j, column=10)
    cell.value = 'Пользователь'
    cell = sheet.cell(row=j, column=11)
    cell.value = 'Поставщик'
    cell = sheet.cell(row=j, column=12)
    cell.value = 'Комментарий'

    for i in range(len(query)):
        j += 1
        data = str(query[i].date)
        material = query[i].material.name
        mark = query[i].mark.name
        object = query[i].object.name
        block1 = ''
        for k in range(len(query[i].block.all())):
            block1 = block1 + str(query[i].block.all()[k].name) + ', '
        block = str(block1)[:-2]
        floor1 = ''
        for k in range(len(query[i].floor.all())):
            floor1 = floor1 + str(query[i].floor.all()[k].name) + ', '
        floor = str(floor1)[:-2]
        try:
            constructive = query[i].constructive.name
        except:
            constructive = ' '
        unit = query[i].unit.name
        count = query[i].count
        user = query[i].user.last_name + " " + query[i].user.first_name
        provider = query[i].provider.name
        comments = query[i].comments
        cell = sheet.cell(row=j, column=1)
        cell.value = data
        cell = sheet.cell(row=j, column=2)
        cell.value = material
        cell = sheet.cell(row=j, column=3)
        cell.value = mark
        cell = sheet.cell(row=j, column=4)
        cell.value = object
        cell = sheet.cell(row=j, column=5)
        cell.value = block
        cell = sheet.cell(row=j, column=6)
        cell.value = floor
        cell = sheet.cell(row=j, column=7)
        cell.value = constructive
        cell = sheet.cell(row=j, column=8)
        cell.value = unit
        cell = sheet.cell(row=j, column=9)
        cell.value = count
        cell = sheet.cell(row=j, column=10)
        cell.value = user
        cell = sheet.cell(row=j, column=11)
        cell.value = provider
        cell = sheet.cell(row=j, column=12)
        cell.value = comments
    file_name = 'file.xlsx'
    wb.save(file_name)
    return wb
